import json
from pathlib import Path

from openpyxl import Workbook

from job_application_agent.cli import QUALITY_GATE_FAILURE_EXIT_CODE, main


def test_tracker_headers_cli_prints_log_safe_headers(
    tmp_path: Path, capsys
) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company", "Role", "Status"])
    applications.append(["Example Co", "Engineering Manager", "Applied"])
    workbook.save(tracker_path)

    env_file = _write_env_file(tmp_path, tracker_path)

    exit_code = main(
        [
            "tracker-headers",
            "--env-file",
            str(env_file),
            "--tab",
            "Applications",
        ]
    )

    assert exit_code == 0
    output = json.loads(capsys.readouterr().out)
    assert output["config"]["has_staging_tracker_path"] is True
    assert output["tracker_headers"]["tabs"][0]["title"] == "Applications"
    assert output["tracker_headers"]["tabs"][0]["headers"] == [
        "Company",
        "Role",
        "Status",
    ]
    assert "Example Co" not in repr(output)
    assert str(tracker_path) not in repr(output)


def test_tracker_schema_cli_prints_log_safe_mapping(
    tmp_path: Path, capsys
) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company", "Role", "Status", "Custom Notes"])
    applications.append(["Example Co", "Engineering Manager", "Applied", "private note"])
    workbook.save(tracker_path)

    env_file = _write_env_file(tmp_path, tracker_path)

    exit_code = main(
        [
            "tracker-schema",
            "--env-file",
            str(env_file),
            "--tab",
            "Applications",
        ]
    )

    assert exit_code == 0
    output = json.loads(capsys.readouterr().out)
    tab = output["tracker_schema"]["tabs"][0]
    assert tab["title"] == "Applications"
    assert tab["is_complete"] is True
    assert tab["mapped_fields"][0]["canonical_field"] == "company"
    assert tab["mapped_fields"][1]["canonical_field"] == "role"
    assert tab["mapped_fields"][2]["canonical_field"] == "status"
    assert tab["unmapped_headers"] == ["Custom Notes"]
    assert "Example Co" not in repr(output)
    assert "private note" not in repr(output)
    assert str(tracker_path) not in repr(output)


def test_tracker_quality_cli_prints_log_safe_counts(
    tmp_path: Path, capsys
) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company", "Role", "Status"])
    applications.append(["Example Co", "Engineering Manager", "Applied"])
    applications.append([None, "Architect", "Applied"])
    workbook.save(tracker_path)

    env_file = _write_env_file(tmp_path, tracker_path)

    exit_code = main(
        [
            "tracker-quality",
            "--env-file",
            str(env_file),
            "--tab",
            "Applications",
            "--max-records",
            "10",
        ]
    )

    assert exit_code == 0
    output = json.loads(capsys.readouterr().out)
    tab = output["tracker_quality"]["tabs"][0]
    assert output["tracker_quality"]["scanned_records"] == 2
    assert output["tracker_quality"]["has_required_blanks"] is True
    assert output["tracker_quality"]["failed_quality_gates"] == []
    assert tab["title"] == "Applications"
    assert tab["scanned_records"] == 2
    assert tab["has_required_blanks"] is True
    company_quality = next(
        field for field in tab["required_fields"] if field["canonical_field"] == "company"
    )
    assert company_quality["populated_count"] == 1
    assert company_quality["blank_count"] == 1
    assert "Example Co" not in repr(output)
    assert "Engineering Manager" not in repr(output)
    assert str(tracker_path) not in repr(output)


def test_tracker_quality_cli_fails_selected_gate_without_values(
    tmp_path: Path, capsys
) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company", "Role", "Status"])
    applications.append([None, "Engineering Manager", "Applied"])
    workbook.save(tracker_path)

    env_file = _write_env_file(tmp_path, tracker_path)

    exit_code = main(
        [
            "tracker-quality",
            "--env-file",
            str(env_file),
            "--tab",
            "Applications",
            "--fail-on-required-blanks",
        ]
    )

    assert exit_code == QUALITY_GATE_FAILURE_EXIT_CODE
    output = json.loads(capsys.readouterr().out)
    assert output["tracker_quality"]["failed_quality_gates"] == ["required_blanks"]
    assert "Engineering Manager" not in repr(output)
    assert str(tracker_path) not in repr(output)


def test_tracker_quality_cli_strict_fails_all_selected_gates(
    tmp_path: Path, capsys
) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company"])
    applications.append(["Example Co"])
    workbook.save(tracker_path)

    env_file = _write_env_file(tmp_path, tracker_path)

    exit_code = main(
        [
            "tracker-quality",
            "--env-file",
            str(env_file),
            "--tab",
            "Applications",
            "--strict",
        ]
    )

    assert exit_code == QUALITY_GATE_FAILURE_EXIT_CODE
    output = json.loads(capsys.readouterr().out)
    assert output["tracker_quality"]["failed_quality_gates"] == [
        "incomplete_schema",
        "required_blanks",
    ]
    assert "Example Co" not in repr(output)
    assert str(tracker_path) not in repr(output)


def test_mock_score_cli_prints_sanitized_score_reports(tmp_path: Path, capsys) -> None:
    fixture_path = tmp_path / "mock_jobs.json"
    fixture_path.write_text(
        json.dumps(
            [
                {
                    "id": "job_mock_001",
                    "source": "mock",
                    "company": "Example Systems",
                    "title": "Engineering Manager",
                    "location": "Remote, United States",
                    "work_arrangement": "remote",
                    "employment_type": "full-time",
                    "compensation_max": 180000,
                    "posting_url": "https://example.com/jobs/engineering-manager",
                    "requirements": ["Lead engineering teams"],
                    "responsibilities": ["Improve delivery systems"],
                }
            ]
        ),
        encoding="utf-8",
    )

    exit_code = main(["mock-score", "--fixture", str(fixture_path)])

    assert exit_code == 0
    output = json.loads(capsys.readouterr().out)
    assert output["mock_score_reports"]["count"] == 1
    job = output["mock_score_reports"]["jobs"][0]
    assert job["id"] == "job_mock_001"
    assert job["company"] == "Example Systems"
    assert job["score"] >= 0
    assert job["rules"]
    assert str(fixture_path) not in repr(output)


def test_mock_queue_cli_filters_by_min_score_and_priority(
    tmp_path: Path, capsys
) -> None:
    fixture_path = _write_mock_queue_fixture(tmp_path)

    exit_code = main(
        [
            "mock-queue",
            "--fixture",
            str(fixture_path),
            "--min-score",
            "50",
            "--priority",
            "medium",
        ]
    )

    assert exit_code == 0
    output = json.loads(capsys.readouterr().out)
    queue = output["mock_queue"]
    assert queue["filters"] == {"min_score": 50, "priorities": ["medium"]}
    assert queue["count"] == 1
    assert queue["items"][0]["job_id"] == "job_mock_medium"
    assert queue["items"][0]["rank"] == 1
    assert queue["items"][0]["priority"] == "medium"
    assert str(fixture_path) not in repr(output)


def test_mock_queue_cli_accepts_repeated_priority_filters_and_reranks(
    tmp_path: Path, capsys
) -> None:
    fixture_path = _write_mock_queue_fixture(tmp_path)

    exit_code = main(
        [
            "mock-queue",
            "--fixture",
            str(fixture_path),
            "--priority",
            "high",
            "--priority",
            "medium",
        ]
    )

    assert exit_code == 0
    output = json.loads(capsys.readouterr().out)
    queue = output["mock_queue"]
    assert queue["filters"] == {"min_score": None, "priorities": ["high", "medium"]}
    assert queue["count"] == 2
    assert [item["job_id"] for item in queue["items"]] == [
        "job_mock_high",
        "job_mock_medium",
    ]
    assert [item["rank"] for item in queue["items"]] == [1, 2]
    assert str(fixture_path) not in repr(output)


def _write_env_file(tmp_path: Path, tracker_path: Path) -> Path:
    env_file = tmp_path / ".env"
    env_file.write_text(
        "\n".join(
            [
                "RUNTIME_MODE=staging",
                f"STAGING_TRACKER_PATH={tracker_path}",
            ]
        ),
        encoding="utf-8",
    )
    return env_file


def _write_mock_queue_fixture(tmp_path: Path) -> Path:
    fixture_path = tmp_path / "mock_jobs.json"
    fixture_path.write_text(
        json.dumps(
            [
                {
                    "id": "job_mock_low",
                    "source": "mock",
                    "company": "Example Retail",
                    "title": "Retail Associate",
                    "location": "Onsite, United States",
                    "work_arrangement": "onsite",
                    "employment_type": "part-time",
                    "compensation_max": 50000,
                    "posting_url": "https://example.com/jobs/retail-associate",
                    "requirements": ["Provide customer support"],
                    "responsibilities": ["Maintain public demo inventory"],
                },
                {
                    "id": "job_mock_high",
                    "source": "mock",
                    "company": "Example Systems",
                    "title": "Engineering Manager",
                    "location": "Remote, United States",
                    "work_arrangement": "remote",
                    "employment_type": "full-time",
                    "compensation_max": 180000,
                    "posting_url": "https://example.com/jobs/engineering-manager",
                    "requirements": ["Lead engineering teams"],
                    "responsibilities": ["Improve delivery systems"],
                },
                {
                    "id": "job_mock_medium",
                    "source": "mock",
                    "company": "Example Studio",
                    "title": "Engineering Manager",
                    "location": "United States",
                    "work_arrangement": "unknown",
                    "employment_type": "unknown",
                    "posting_url": "https://example.com/jobs/mock-manager",
                    "requirements": ["Coordinate public-safe planning"],
                    "responsibilities": ["Keep review notes sanitized"],
                },
            ]
        ),
        encoding="utf-8",
    )
    return fixture_path
