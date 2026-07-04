from pathlib import Path

import pytest
from openpyxl import Workbook

from job_application_agent.config import ConfigurationError, RuntimeConfig
from job_application_agent.tracker_quality import LocalTrackerQualityAnalyzer


def test_quality_summary_counts_required_field_blanks_without_values(
    tmp_path: Path,
) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company", "Role", "Status", "Custom Notes"])
    applications.append(["Example Co", "Engineering Manager", "Applied", "sample note"])
    applications.append([None, "Architect", "Applied", "another sample"])
    applications.append(["", "", "", ""])
    workbook.save(tracker_path)

    summary = LocalTrackerQualityAnalyzer(workbook_path=tracker_path).summarize(
        ["Applications"]
    )

    tab = summary.tabs[0]
    assert summary.scanned_records == 2
    assert summary.has_required_blanks is True
    assert tab.scanned_records == 2
    assert tab.blank_records_skipped == 1
    assert tab.is_schema_complete is True
    assert tab.has_required_blanks is True
    assert tab.unmapped_headers == ("Custom Notes",)
    assert tab.missing_required_fields == ()
    quality_by_field = {field.canonical_field: field for field in tab.required_fields}
    assert quality_by_field["company"].populated_count == 1
    assert quality_by_field["company"].blank_count == 1
    assert quality_by_field["company"].has_blanks is True
    assert quality_by_field["role"].populated_count == 2
    assert quality_by_field["role"].blank_count == 0
    assert quality_by_field["status"].populated_count == 2
    assert quality_by_field["status"].blank_count == 0
    assert summary.failed_quality_gates(fail_on_required_blanks=True) == (
        "required_blanks",
    )
    assert "Example Co" not in repr(summary)
    assert "sample note" not in repr(summary)
    assert str(tracker_path) not in repr(summary)


def test_quality_summary_reports_missing_schema_fields_as_blank_counts(
    tmp_path: Path,
) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company"])
    applications.append(["Example Co"])
    workbook.save(tracker_path)

    summary = LocalTrackerQualityAnalyzer(workbook_path=tracker_path).summarize(
        ["Applications"]
    )

    tab = summary.tabs[0]
    assert summary.is_schema_complete is False
    assert summary.has_required_blanks is True
    assert tab.is_schema_complete is False
    assert tab.has_required_blanks is True
    assert tab.missing_required_fields == ("role", "status")
    assert tab.scanned_records == 1
    quality_by_field = {field.canonical_field: field for field in tab.required_fields}
    assert quality_by_field["company"].populated_count == 1
    assert quality_by_field["company"].blank_count == 0
    assert quality_by_field["role"].populated_count == 0
    assert quality_by_field["role"].blank_count == 1
    assert quality_by_field["status"].populated_count == 0
    assert quality_by_field["status"].blank_count == 1
    assert summary.failed_quality_gates(
        fail_on_incomplete_schema=True,
        fail_on_required_blanks=True,
    ) == ("incomplete_schema", "required_blanks")
    assert "Example Co" not in repr(summary)
    assert str(tracker_path) not in repr(summary)


def test_quality_summary_respects_max_records(tmp_path: Path) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company", "Role", "Status"])
    applications.append(["A", "Role A", "Applied"])
    applications.append(["B", "Role B", "Applied"])
    workbook.save(tracker_path)

    summary = LocalTrackerQualityAnalyzer(workbook_path=tracker_path).summarize(
        ["Applications"], max_records=1
    )

    assert summary.tabs[0].scanned_records == 1
    assert summary.tabs[0].truncated is True


def test_quality_summary_passes_quality_gates_for_complete_required_fields(
    tmp_path: Path,
) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    applications = workbook.create_sheet("Applications")
    applications.append(["Company", "Role", "Status"])
    applications.append(["Example Co", "Engineering Manager", "Applied"])
    workbook.save(tracker_path)

    summary = LocalTrackerQualityAnalyzer(workbook_path=tracker_path).summarize(
        ["Applications"]
    )

    assert summary.is_schema_complete is True
    assert summary.has_required_blanks is False
    assert summary.failed_quality_gates(
        fail_on_incomplete_schema=True,
        fail_on_required_blanks=True,
    ) == ()


def test_quality_summary_requires_positive_max_records(tmp_path: Path) -> None:
    config = RuntimeConfig.from_mapping(
        {
            "RUNTIME_MODE": "staging",
            "STAGING_TRACKER_PATH": str(tmp_path / "missing.xlsx"),
        }
    )

    with pytest.raises(ConfigurationError, match="max_records"):
        LocalTrackerQualityAnalyzer.from_config(config).summarize(max_records=0)
