from pathlib import Path
from typing import Any

from openpyxl import Workbook

from job_application_agent.config import RuntimeConfig
from job_application_agent.integrations.sheets import (
    GoogleSheetsAdapter,
    LocalExcelTrackerAdapter,
    MockSheetsAdapter,
    build_sheets_adapter,
    spreadsheet_summary_from_google_payload,
)


EXPECTED_TABS = {
    "Dashboard",
    "Applications",
    "Daily Leads",
    "Contacts",
    "Scan Setup",
    "Lists",
    "Rejected Applications",
    "Match Score Rubric",
    "Recruiter Apply",
}


def test_mock_adapter_returns_expected_tracker_tabs() -> None:
    summary = MockSheetsAdapter().get_metadata()

    assert summary.title == "Mock Job Tracker"
    assert set(summary.tab_titles) == EXPECTED_TABS


def test_build_sheets_adapter_defaults_to_mock() -> None:
    adapter = build_sheets_adapter(RuntimeConfig.from_mapping({}))

    assert isinstance(adapter, MockSheetsAdapter)


def test_build_sheets_adapter_uses_local_excel_for_staging_when_path_exists() -> None:
    config = RuntimeConfig.from_mapping(
        {
            "RUNTIME_MODE": "staging",
            "STAGING_TRACKER_PATH": "C:/private/staging_job_tracker.xlsx",
        }
    )

    adapter = build_sheets_adapter(config)

    assert isinstance(adapter, LocalExcelTrackerAdapter)


def test_build_sheets_adapter_uses_google_for_staging_without_local_path() -> None:
    config = RuntimeConfig.from_mapping(
        {
            "RUNTIME_MODE": "staging",
            "GOOGLE_SHEETS_STAGING_ID": "private-id",
        }
    )

    adapter = build_sheets_adapter(config)

    assert isinstance(adapter, GoogleSheetsAdapter)


def test_local_excel_adapter_reads_workbook_metadata(tmp_path: Path) -> None:
    tracker_path = tmp_path / "staging_job_tracker.xlsx"
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard.freeze_panes = "A2"
    dashboard.append(["Metric", "Value"])
    dashboard.append(["Applications", 3])
    applications = workbook.create_sheet("Applications")
    applications.append(["Company", "Role"])
    applications.append(["Example Co", "Engineering Manager"])
    workbook.create_sheet("Daily Leads")
    workbook.save(tracker_path)

    summary = LocalExcelTrackerAdapter(workbook_path=tracker_path).get_metadata()

    assert summary.title == "staging_job_tracker"
    assert summary.tab_titles == ("Dashboard", "Applications", "Daily Leads")
    assert summary.tabs[0].row_count == 2
    assert summary.tabs[0].column_count == 2
    assert summary.tabs[0].frozen_row_count == 1
    assert str(tracker_path) not in repr(summary)


def test_spreadsheet_summary_from_google_payload_is_log_safe() -> None:
    payload = {
        "spreadsheetId": "private-id-that-should-not-be-returned",
        "properties": {
            "title": "Copy of Job Tracker",
            "locale": "en_US",
            "timeZone": "America/New_York",
        },
        "sheets": [
            {
                "properties": {
                    "sheetId": 101,
                    "title": "Applications",
                    "index": 1,
                    "gridProperties": {
                        "rowCount": 93,
                        "columnCount": 31,
                        "frozenRowCount": 1,
                    },
                }
            }
        ],
    }

    summary = spreadsheet_summary_from_google_payload(payload)

    assert summary.title == "Copy of Job Tracker"
    assert summary.time_zone == "America/New_York"
    assert summary.tabs[0].title == "Applications"
    assert summary.tabs[0].row_count == 93
    assert "private-id-that-should-not-be-returned" not in repr(summary)


def test_google_adapter_reads_metadata_with_private_id_only_in_request() -> None:
    private_id = "private-id"
    fake_service = FakeSheetsService(
        {
            "properties": {"title": "Copy of Job Tracker"},
            "sheets": [
                {
                    "properties": {
                        "sheetId": 0,
                        "title": "Dashboard",
                        "index": 0,
                        "gridProperties": {"rowCount": 100, "columnCount": 26},
                    }
                }
            ],
        }
    )
    adapter = GoogleSheetsAdapter(spreadsheet_id=private_id, service=fake_service)

    summary = adapter.get_metadata()

    assert summary.title == "Copy of Job Tracker"
    assert summary.tab_titles == ("Dashboard",)
    assert fake_service.last_get_kwargs["spreadsheetId"] == private_id
    assert fake_service.last_get_kwargs["fields"] == GoogleSheetsAdapter.METADATA_FIELDS
    assert private_id not in repr(summary)


class FakeSheetsService:
    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload
        self.last_get_kwargs: dict[str, Any] = {}

    def spreadsheets(self) -> "FakeSheetsService":
        return self

    def get(self, **kwargs: Any) -> "FakeRequest":
        self.last_get_kwargs = kwargs
        return FakeRequest(self.payload)


class FakeRequest:
    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload

    def execute(self) -> dict[str, Any]:
        return self.payload
