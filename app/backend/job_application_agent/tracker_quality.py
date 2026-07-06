"""Public-safe tracker quality summary helpers.

This module scans local tracker worksheets to produce aggregate quality counts.
It does not return individual row values and does not write to the tracker.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any, Sequence
from urllib.parse import urlparse

from job_application_agent.config import ConfigurationError, RuntimeConfig
from job_application_agent.integrations.sheets import (
    DEFAULT_TRACKER_HEADER_TABS,
    LocalExcelTrackerAdapter,
    SheetsAdapterError,
)
from job_application_agent.tracker_schema import (
    HeaderFieldMapping,
    TabSchemaMapping,
    TRACKER_TAB_SCHEMAS,
    map_tracker_headers,
)

FORMAT_FIELD_KINDS: dict[str, str] = {
    "contact_date": "date",
    "date_applied": "date",
    "email": "email",
    "job_url": "url",
    "last_contacted": "date",
    "match_score": "score",
    "status": "status",
}

ALLOWED_STATUS_VALUES = frozenset(
    {
        "accepted",
        "applied",
        "archived",
        "closed",
        "contacted",
        "declined",
        "follow up",
        "followup",
        "in progress",
        "in review",
        "interview",
        "interviewing",
        "lead",
        "new",
        "not started",
        "offer",
        "onsite",
        "phone screen",
        "queued",
        "rejected",
        "review",
        "saved",
        "screen",
        "submitted",
        "withdrawn",
    }
)


@dataclass(frozen=True)
class RequiredFieldQuality:
    """Aggregate quality counts for one required canonical field."""

    canonical_field: str
    populated_count: int
    blank_count: int

    @property
    def has_blanks(self) -> bool:
        return self.blank_count > 0


@dataclass(frozen=True)
class FieldFormatQuality:
    """Aggregate format counts for one mapped canonical field."""

    canonical_field: str
    value_kind: str
    checked_count: int
    blank_count: int
    invalid_count: int

    @property
    def has_invalid_values(self) -> bool:
        return self.invalid_count > 0


@dataclass(frozen=True)
class TabQualitySummary:
    """Public-safe quality summary for one tracker tab."""

    title: str
    scanned_records: int
    blank_records_skipped: int
    is_schema_complete: bool
    missing_required_fields: tuple[str, ...]
    unmapped_headers: tuple[str, ...]
    required_fields: tuple[RequiredFieldQuality, ...]
    format_fields: tuple[FieldFormatQuality, ...]
    truncated: bool

    @property
    def has_required_blanks(self) -> bool:
        return any(field.has_blanks for field in self.required_fields)

    @property
    def has_format_warnings(self) -> bool:
        return any(field.has_invalid_values for field in self.format_fields)


@dataclass(frozen=True)
class TrackerQualitySummary:
    """Public-safe quality summary for tracker tabs."""

    tabs: tuple[TabQualitySummary, ...]
    max_records: int

    @property
    def scanned_records(self) -> int:
        return sum(tab.scanned_records for tab in self.tabs)

    @property
    def is_schema_complete(self) -> bool:
        return all(tab.is_schema_complete for tab in self.tabs)

    @property
    def has_required_blanks(self) -> bool:
        return any(tab.has_required_blanks for tab in self.tabs)

    @property
    def has_format_warnings(self) -> bool:
        return any(tab.has_format_warnings for tab in self.tabs)

    def failed_quality_gates(
        self,
        *,
        fail_on_incomplete_schema: bool = False,
        fail_on_required_blanks: bool = False,
        fail_on_format_warnings: bool = False,
    ) -> tuple[str, ...]:
        failures: list[str] = []
        if fail_on_incomplete_schema and not self.is_schema_complete:
            failures.append("incomplete_schema")
        if fail_on_required_blanks and self.has_required_blanks:
            failures.append("required_blanks")
        if fail_on_format_warnings and self.has_format_warnings:
            failures.append("format_warnings")
        return tuple(failures)


class LocalTrackerQualityAnalyzer:
    """Read-only local Excel tracker quality analyzer."""

    def __init__(self, *, workbook_path: str | Path) -> None:
        self._workbook_path = Path(workbook_path)
        self._metadata_adapter = LocalExcelTrackerAdapter(
            workbook_path=self._workbook_path
        )

    @classmethod
    def from_config(cls, config: RuntimeConfig) -> "LocalTrackerQualityAnalyzer":
        if not config.staging_tracker_path:
            raise ConfigurationError(
                "STAGING_TRACKER_PATH is required for local tracker quality summaries."
            )
        return cls(workbook_path=config.staging_tracker_path)

    def summarize(
        self,
        tab_titles: Sequence[str] | None = None,
        *,
        header_row: int = 1,
        max_records: int = 1000,
    ) -> TrackerQualitySummary:
        if max_records < 1:
            raise ConfigurationError("max_records must be 1 or greater.")

        selected_titles = tuple(tab_titles or DEFAULT_TRACKER_HEADER_TABS)
        headers = self._metadata_adapter.get_headers(
            selected_titles, header_row=header_row
        )
        schema_mapping = map_tracker_headers(headers)
        workbook = _load_workbook(self._workbook_path)
        try:
            worksheets_by_title = {
                worksheet.title: worksheet for worksheet in workbook.worksheets
            }
            tab_summaries = tuple(
                _summarize_tab_quality(
                    worksheets_by_title[tab_mapping.title],
                    tab_mapping,
                    header_row=header_row,
                    max_records=max_records,
                )
                for tab_mapping in schema_mapping.tabs
            )
        finally:
            workbook.close()

        return TrackerQualitySummary(tabs=tab_summaries, max_records=max_records)


def _summarize_tab_quality(
    worksheet: Any,
    tab_mapping: TabSchemaMapping,
    *,
    header_row: int,
    max_records: int,
) -> TabQualitySummary:
    required_fields = tuple(
        field.name
        for field in TRACKER_TAB_SCHEMAS.get(tab_mapping.title, ())
        if field.required
    )
    required_columns = _required_column_indexes(tab_mapping.mapped_fields, required_fields)
    format_columns = _format_column_indexes(tab_mapping.mapped_fields)
    populated_counts = {field_name: 0 for field_name in required_fields}
    blank_counts = {field_name: 0 for field_name in required_fields}
    format_counts = _initial_format_counts(format_columns)
    scanned_records = 0
    blank_records_skipped = 0
    truncated = False

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if _is_blank_record(row):
            blank_records_skipped += 1
            continue
        if scanned_records >= max_records:
            truncated = True
            break
        scanned_records += 1
        for field_name, column_index in required_columns.items():
            value = _value_at(row, column_index)
            if _is_blank_value(value):
                blank_counts[field_name] += 1
            else:
                populated_counts[field_name] += 1
        _count_format_fields(row, format_columns, format_counts)

    _count_missing_required_fields_as_blank(
        required_fields,
        required_columns,
        blank_counts,
        scanned_records=scanned_records,
    )

    return TabQualitySummary(
        title=tab_mapping.title,
        scanned_records=scanned_records,
        blank_records_skipped=blank_records_skipped,
        is_schema_complete=tab_mapping.is_complete,
        missing_required_fields=tab_mapping.missing_required_fields,
        unmapped_headers=tab_mapping.unmapped_headers,
        required_fields=tuple(
            RequiredFieldQuality(
                canonical_field=field_name,
                populated_count=populated_counts[field_name],
                blank_count=blank_counts[field_name],
            )
            for field_name in required_fields
        ),
        format_fields=tuple(
            FieldFormatQuality(
                canonical_field=field_name,
                value_kind=FORMAT_FIELD_KINDS[field_name],
                checked_count=format_counts[field_name]["checked_count"],
                blank_count=format_counts[field_name]["blank_count"],
                invalid_count=format_counts[field_name]["invalid_count"],
            )
            for field_name in sorted(format_columns)
        ),
        truncated=truncated,
    )


def _required_column_indexes(
    mapped_fields: tuple[HeaderFieldMapping, ...],
    required_fields: tuple[str, ...],
) -> dict[str, int]:
    required_field_set = set(required_fields)
    return {
        field.canonical_field: field.column_index
        for field in mapped_fields
        if field.canonical_field in required_field_set
    }


def _format_column_indexes(
    mapped_fields: tuple[HeaderFieldMapping, ...]
) -> dict[str, int]:
    return {
        field.canonical_field: field.column_index
        for field in mapped_fields
        if field.canonical_field in FORMAT_FIELD_KINDS
    }


def _initial_format_counts(
    format_columns: dict[str, int]
) -> dict[str, dict[str, int]]:
    return {
        field_name: {"blank_count": 0, "checked_count": 0, "invalid_count": 0}
        for field_name in format_columns
    }


def _count_format_fields(
    row: tuple[Any, ...],
    format_columns: dict[str, int],
    format_counts: dict[str, dict[str, int]],
) -> None:
    for field_name, column_index in format_columns.items():
        value = _value_at(row, column_index)
        if _is_blank_value(value):
            format_counts[field_name]["blank_count"] += 1
            continue
        format_counts[field_name]["checked_count"] += 1
        if not _has_valid_format(field_name, value):
            format_counts[field_name]["invalid_count"] += 1


def _count_missing_required_fields_as_blank(
    required_fields: tuple[str, ...],
    required_columns: dict[str, int],
    blank_counts: dict[str, int],
    *,
    scanned_records: int,
) -> None:
    for field_name in required_fields:
        if field_name not in required_columns:
            blank_counts[field_name] = scanned_records


def _load_workbook(path: Path) -> Any:
    if not path.exists():
        raise SheetsAdapterError(
            "Local staging tracker export was not found. Confirm STAGING_TRACKER_PATH."
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - import depends on optional deps
        raise SheetsAdapterError(
            "Local Excel support requires openpyxl. "
            "Install with: pip install -e '.[excel]'"
        ) from exc

    try:
        return load_workbook(filename=path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - depends on file corruption details
        raise SheetsAdapterError(
            "Failed to read local staging tracker export. Confirm it is a valid .xlsx file."
        ) from exc


def _value_at(row: tuple[Any, ...], column_index: int) -> Any:
    zero_based_index = column_index - 1
    if zero_based_index >= len(row):
        return None
    return row[zero_based_index]


def _is_blank_record(row: tuple[Any, ...]) -> bool:
    return all(_is_blank_value(value) for value in row)


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    return str(value).strip() == ""


def _has_valid_format(field_name: str, value: Any) -> bool:
    value_kind = FORMAT_FIELD_KINDS[field_name]
    if value_kind == "date":
        return _has_valid_date_format(value)
    if value_kind == "email":
        return _has_valid_email_format(value)
    if value_kind == "score":
        return _has_valid_score_format(value)
    if value_kind == "status":
        return _has_valid_status_format(value)
    if value_kind == "url":
        return _has_valid_url_format(value)
    return True


def _has_valid_date_format(value: Any) -> bool:
    if isinstance(value, datetime | date):
        return True
    if not isinstance(value, str):
        return False
    stripped_value = value.strip()
    if not stripped_value:
        return False
    for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            datetime.strptime(stripped_value, date_format)
            return True
        except ValueError:
            continue
    return False


def _has_valid_email_format(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value.strip()))


def _has_valid_score_format(value: Any) -> bool:
    try:
        score = float(str(value).strip().removesuffix("%"))
    except ValueError:
        return False
    return 0 <= score <= 100


def _has_valid_status_format(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    return _normalize_status_value(value) in ALLOWED_STATUS_VALUES


def _has_valid_url_format(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    parsed_url = urlparse(value.strip())
    return parsed_url.scheme in {"http", "https"} and bool(parsed_url.netloc)


def _normalize_status_value(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.strip().lower()).strip()
