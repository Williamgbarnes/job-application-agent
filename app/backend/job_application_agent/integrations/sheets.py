"""Read-only tracker metadata adapters.

The adapters intentionally start with spreadsheet metadata, tab discovery, and
header discovery only. They do not expose write methods and do not hard-code
private spreadsheet IDs or local private file paths.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Protocol, Sequence

from job_application_agent.config import ConfigurationError, RuntimeConfig

DEFAULT_TRACKER_HEADER_TABS = (
    "Applications",
    "Daily Leads",
    "Contacts",
    "Rejected Applications",
    "Recruiter Apply",
)


class SheetsAdapterError(RuntimeError):
    """Raised when a tracker adapter cannot complete a read-only operation."""


@dataclass(frozen=True)
class SheetTab:
    """Public-safe summary of one spreadsheet tab."""

    title: str
    sheet_id: int
    index: int
    row_count: int | None = None
    column_count: int | None = None
    frozen_row_count: int | None = None


@dataclass(frozen=True)
class SpreadsheetSummary:
    """Public-safe spreadsheet metadata.

    Private IDs and local file paths are deliberately excluded so summaries are
    safe to log.
    """

    title: str
    locale: str | None
    time_zone: str | None
    tabs: tuple[SheetTab, ...]

    @property
    def tab_titles(self) -> tuple[str, ...]:
        return tuple(tab.title for tab in self.tabs)


@dataclass(frozen=True)
class SheetHeaders:
    """Public-safe header summary for one tracker tab."""

    title: str
    index: int
    header_row: int
    headers: tuple[str, ...]


@dataclass(frozen=True)
class TrackerHeadersSummary:
    """Public-safe tracker header summary.

    Cell values below the header row are deliberately excluded.
    """

    tabs: tuple[SheetHeaders, ...]

    @property
    def tab_titles(self) -> tuple[str, ...]:
        return tuple(tab.title for tab in self.tabs)


class SheetsMetadataProvider(Protocol):
    """Protocol shared by mock, local-file, and Google-backed adapters."""

    def get_metadata(self) -> SpreadsheetSummary:
        """Return read-only spreadsheet metadata."""

    def get_headers(
        self,
        tab_titles: Sequence[str] | None = None,
        *,
        header_row: int = 1,
    ) -> TrackerHeadersSummary:
        """Return read-only tracker header metadata."""


class MockSheetsAdapter:
    """Mock adapter for public demos and tests."""

    MOCK_HEADERS: Mapping[str, tuple[str, ...]] = {
        "Applications": ("Company", "Role", "Status", "Date Applied"),
        "Daily Leads": ("Company", "Role", "URL", "Match Score"),
        "Contacts": ("Name", "Company", "Email", "Last Contacted"),
        "Rejected Applications": ("Company", "Role", "Status", "Archive Reason"),
        "Recruiter Apply": ("Company", "Role", "Recruiter", "Next Step"),
    }

    def __init__(self, summary: SpreadsheetSummary | None = None) -> None:
        self._summary = summary or SpreadsheetSummary(
            title="Mock Job Tracker",
            locale="en_US",
            time_zone="America/New_York",
            tabs=(
                SheetTab(title="Dashboard", sheet_id=0, index=0),
                SheetTab(title="Applications", sheet_id=101, index=1),
                SheetTab(title="Daily Leads", sheet_id=102, index=2),
                SheetTab(title="Contacts", sheet_id=103, index=3),
                SheetTab(title="Scan Setup", sheet_id=104, index=4),
                SheetTab(title="Lists", sheet_id=105, index=5),
                SheetTab(title="Rejected Applications", sheet_id=106, index=6),
                SheetTab(title="Match Score Rubric", sheet_id=107, index=7),
                SheetTab(title="Recruiter Apply", sheet_id=108, index=8),
            ),
        )

    def get_metadata(self) -> SpreadsheetSummary:
        return self._summary

    def get_headers(
        self,
        tab_titles: Sequence[str] | None = None,
        *,
        header_row: int = 1,
    ) -> TrackerHeadersSummary:
        selected_titles = tuple(tab_titles or DEFAULT_TRACKER_HEADER_TABS)
        metadata_by_title = {tab.title: tab for tab in self._summary.tabs}
        return TrackerHeadersSummary(
            tabs=tuple(
                SheetHeaders(
                    title=title,
                    index=metadata_by_title.get(title, SheetTab(title, 0, 0)).index,
                    header_row=header_row,
                    headers=self.MOCK_HEADERS.get(title, ()),
                )
                for title in selected_titles
            )
        )


class LocalExcelTrackerAdapter:
    """Read-only adapter for a private local `.xlsx` tracker export."""

    def __init__(self, *, workbook_path: str | Path) -> None:
        path = Path(workbook_path)
        if not path:
            raise ConfigurationError("A workbook path is required.")
        self._workbook_path = path

    @classmethod
    def from_config(cls, config: RuntimeConfig) -> "LocalExcelTrackerAdapter":
        if not config.staging_tracker_path:
            raise ConfigurationError("STAGING_TRACKER_PATH is required for local tracker reads.")
        return cls(workbook_path=config.staging_tracker_path)

    def get_metadata(self) -> SpreadsheetSummary:
        workbook = self._load_workbook()
        try:
            tabs = tuple(
                SheetTab(
                    title=worksheet.title,
                    sheet_id=index,
                    index=index,
                    row_count=worksheet.max_row,
                    column_count=worksheet.max_column,
                    frozen_row_count=_frozen_row_count(worksheet.freeze_panes),
                )
                for index, worksheet in enumerate(workbook.worksheets)
            )
        finally:
            workbook.close()

        return SpreadsheetSummary(
            title=self._workbook_path.stem,
            locale=None,
            time_zone=None,
            tabs=tabs,
        )

    def get_headers(
        self,
        tab_titles: Sequence[str] | None = None,
        *,
        header_row: int = 1,
    ) -> TrackerHeadersSummary:
        if header_row < 1:
            raise ConfigurationError("header_row must be 1 or greater.")

        selected_titles = tuple(tab_titles or DEFAULT_TRACKER_HEADER_TABS)
        workbook = self._load_workbook()
        try:
            worksheets_by_title = {worksheet.title: worksheet for worksheet in workbook.worksheets}
            missing_titles = [
                title for title in selected_titles if title not in worksheets_by_title
            ]
            if missing_titles:
                raise SheetsAdapterError(
                    "Local staging tracker is missing expected tabs: "
                    + ", ".join(missing_titles)
                )

            tabs = tuple(
                SheetHeaders(
                    title=title,
                    index=workbook.worksheets.index(worksheets_by_title[title]),
                    header_row=header_row,
                    headers=_read_header_row(worksheets_by_title[title], header_row),
                )
                for title in selected_titles
            )
        finally:
            workbook.close()

        return TrackerHeadersSummary(tabs=tabs)

    def _load_workbook(self) -> Any:
        if not self._workbook_path.exists():
            raise SheetsAdapterError(
                "Local staging tracker export was not found. Confirm STAGING_TRACKER_PATH."
            )

        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - import depends on optional deps
            raise SheetsAdapterError(
                "Local Excel support requires openpyxl. Install with: "
                "pip install -e '.[excel]'"
            ) from exc

        try:
            return load_workbook(
                filename=self._workbook_path, read_only=False, data_only=False
            )
        except Exception as exc:  # pragma: no cover - depends on file corruption details
            raise SheetsAdapterError(
                "Failed to read local staging tracker export. Confirm it is a valid .xlsx file."
            ) from exc


class GoogleSheetsAdapter:
    """Read-only Google Sheets adapter for staging/private runtime use."""

    METADATA_FIELDS = (
        "properties(title,locale,timeZone),"
        "sheets(properties(sheetId,title,index,gridProperties(rowCount,columnCount,frozenRowCount)))"
    )

    def __init__(self, *, spreadsheet_id: str, service: Any | None = None) -> None:
        if not spreadsheet_id:
            raise ConfigurationError("A spreadsheet ID is required for GoogleSheetsAdapter.")
        self._spreadsheet_id = spreadsheet_id
        self._service = service

    @classmethod
    def from_config(cls, config: RuntimeConfig) -> "GoogleSheetsAdapter":
        if config.runtime_mode == "staging":
            spreadsheet_id = config.google_sheets_staging_id
        elif config.runtime_mode == "production":
            spreadsheet_id = config.google_sheets_production_id
        else:
            raise ConfigurationError(
                "GoogleSheetsAdapter requires RUNTIME_MODE=staging or production."
            )

        if not spreadsheet_id:
            raise ConfigurationError(
                "Configured runtime mode is missing its Google Sheets spreadsheet ID."
            )

        return cls(spreadsheet_id=spreadsheet_id)

    def get_metadata(self) -> SpreadsheetSummary:
        service = self._service or _build_google_sheets_service()
        try:
            payload = (
                service.spreadsheets()
                .get(spreadsheetId=self._spreadsheet_id, fields=self.METADATA_FIELDS)
                .execute()
            )
        except Exception as exc:  # pragma: no cover - depends on Google client behavior
            raise SheetsAdapterError(
                "Failed to read Google Sheets metadata. Confirm credentials and access."
            ) from exc

        return spreadsheet_summary_from_google_payload(payload)

    def get_headers(
        self,
        tab_titles: Sequence[str] | None = None,
        *,
        header_row: int = 1,
    ) -> TrackerHeadersSummary:
        raise SheetsAdapterError(
            "Header discovery for live Google Sheets is not implemented yet. "
            "Use STAGING_TRACKER_PATH for local tracker header discovery."
        )


def build_sheets_adapter(config: RuntimeConfig) -> SheetsMetadataProvider:
    """Build the right adapter for the selected runtime mode."""

    if config.runtime_mode == "mock":
        return MockSheetsAdapter()
    if config.runtime_mode == "staging" and config.staging_tracker_path:
        return LocalExcelTrackerAdapter.from_config(config)
    return GoogleSheetsAdapter.from_config(config)


def spreadsheet_summary_from_google_payload(
    payload: Mapping[str, Any]
) -> SpreadsheetSummary:
    """Convert a Google Sheets API spreadsheet payload to a safe summary."""

    properties = payload.get("properties") or {}
    tabs: list[SheetTab] = []

    for sheet in payload.get("sheets") or []:
        sheet_properties = sheet.get("properties") or {}
        grid_properties = sheet_properties.get("gridProperties") or {}
        tabs.append(
            SheetTab(
                title=str(sheet_properties.get("title") or ""),
                sheet_id=int(sheet_properties.get("sheetId") or 0),
                index=int(sheet_properties.get("index") or 0),
                row_count=_optional_int(grid_properties.get("rowCount")),
                column_count=_optional_int(grid_properties.get("columnCount")),
                frozen_row_count=_optional_int(grid_properties.get("frozenRowCount")),
            )
        )

    return SpreadsheetSummary(
        title=str(properties.get("title") or "Untitled spreadsheet"),
        locale=_optional_str(properties.get("locale")),
        time_zone=_optional_str(properties.get("timeZone")),
        tabs=tuple(tabs),
    )


def _build_google_sheets_service() -> Any:
    """Build a Google Sheets service using application default credentials."""

    try:
        import google.auth
        from googleapiclient.discovery import build
    except ImportError as exc:  # pragma: no cover - import depends on optional deps
        raise SheetsAdapterError(
            "Google Sheets support requires the optional 'google' dependencies. "
            "Install with: pip install -e '.[google]'"
        ) from exc

    credentials, _ = google.auth.default(
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
    )
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def _read_header_row(worksheet: Any, header_row: int) -> tuple[str, ...]:
    values = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            values_only=True,
        ),
        (),
    )
    return tuple(_normalize_header(value) for value in values if _normalize_header(value))


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _frozen_row_count(freeze_panes: str | None) -> int | None:
    if not freeze_panes:
        return None
    row_digits = "".join(character for character in freeze_panes if character.isdigit())
    if not row_digits:
        return None
    return max(int(row_digits) - 1, 0)


def _optional_int(value: Any) -> int | None:
    if value is None:
        return None
    return int(value)


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
